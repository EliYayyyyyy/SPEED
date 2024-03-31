import os
import csv
from openpyxl import load_workbook

# Directory containing the CSV files
directory = "/Users/qichenyuan/PRIDICT/batch"
excel_file_path = "/Users/qichenyuan/PRIDICT/batch/List.xlsx"
csv_folder_path = "/Users/qichenyuan/PRIDICT/batch"

# Function to extract specific cell value from CSV
def extract_csv_value(csv_file_path, row, col):
    if os.path.exists(csv_file_path):
        with open(csv_file_path, 'r') as csv_file:
            lines = csv_file.readlines()
            if len(lines) > 1:
                return lines[1].split(',')[col].strip()
    return None

# Function to update Excel with CSV values
def update_excel(excel_file_path, csv_folder_path):
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
        sequence_name = row[0].value
        csv_file_path = os.path.join(csv_folder_path, f"{sequence_name}_nicking_guides.csv")
        b2_value = extract_csv_value(csv_file_path, 1, 1)
        if b2_value:
            cell = sheet.cell(row=row[0].row, column=14)
            cell.value = b2_value

    workbook.save(excel_file_path)

# Function to update Excel with values from another CSV
def update_excel_with_csv(excel_file_path, csv_folder_path):
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active
    empty_row = 2
    while sheet.cell(row=empty_row, column=1).value is not None:
        empty_row += 1
    current_row = 2
    for row in sheet.iter_rows(min_row=current_row, max_col=1, max_row=empty_row - 1, values_only=True):
        sequence_name = row[0]
        csv_file_path = os.path.join(csv_folder_path, f"{sequence_name}_pegRNA_Pridict_full.csv")
        if os.path.exists(csv_file_path):
            with open(csv_file_path, 'r') as csv_file:
                csv_reader = csv.reader(csv_file)
                csv_data = list(csv_reader)
                if len(csv_data) > 1:
                    b_value = csv_data[1][1]
                    c_value = csv_data[1][2]
                    n_value = csv_data[1][13]
                    q_value = csv_data[1][16]
                    o_value = csv_data[1][14]
                    sheet.cell(row=current_row, column=2, value=b_value)
                    sheet.cell(row=current_row, column=3, value=c_value)
                    sheet.cell(row=current_row, column=4, value=n_value)
                    sheet.cell(row=current_row, column=6, value=q_value)
                    sheet.cell(row=current_row, column=7, value=o_value)
                    current_row += 1
    workbook.save(excel_file_path)

# Update Excel with nicking_guides.csv values
update_excel(excel_file_path, csv_folder_path)

# Update Excel with pegRNA_Pridict_full.csv values
update_excel_with_csv(excel_file_path, csv_folder_path)
