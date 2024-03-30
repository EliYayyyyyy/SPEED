from openpyxl import load_workbook
from Bio.Seq import Seq

# Load the Excel file
file_path = "/Users/qichenyuan/Desktop/BootStrap Bio/NAFLD_03302024.xlsx"
workbook = load_workbook(file_path)
sheet = workbook["PRIDICT Results"]

# Assuming column B,D and N contain DNA sequences
sequences_N = [sheet[f'N{row}'].value for row in range(2, sheet.max_row + 1)]
sequences_B = [sheet[f'B{row}'].value for row in range(2, sheet.max_row + 1)]
# Use 2-20bp of the protospacer (remove the first G)
sequences_D = [sheet[f'D{row}'].value[1:20] for row in range(2, sheet.max_row + 1)]

# Function to add one more base from x to the beginning of y if entire y matches x
def find_and_extract(sequence_x, sequence_y):
    # Convert sequences to BioPython Seq objects
    seq_x = Seq(sequence_x)
    seq_y = Seq(sequence_y)

    # Find the index of the first occurrence of sequence_y in sequence_x
    index = seq_x.find(seq_y)

    if index != -1:
        # If sequence D is found in sequence B
        extracted_seq = seq_x[index - 1:index + len(seq_y)]
        return str(extracted_seq)
    else:
        # Search for sequence y in the reverse complement strand of sequence x
        seq_x_rc = seq_x.reverse_complement()
        index_rc = seq_x_rc.find(seq_y)
        if index_rc != -1:
            # Extract the DNA containing 1 base upstream of sequence D and sequence D from the reverse complement strand
            extracted_seq_rc = seq_x_rc[index_rc - 1:index_rc + len(seq_y)]
            return str(extracted_seq_rc)
        else:
            print("Sequence Y not found in sequence X or its reverse complement")
            return None

# find_and_extract all pairs of sequences_B and sequences_N in list as modified_sequences_N
modified_sequences_N = [find_and_extract(sequence_B, sequence_N) for sequence_B, sequence_N in zip(sequences_B, sequences_N)]

# Write the modified nicking guide sequences to the Excel file
for index, sequence in enumerate(modified_sequences_N):
    sheet[f'N{index+2}'].value = sequence

# find_and_extract all pairs of sequences_B and sequences_D in list as modified_sequences_D
modified_sequences_D = [find_and_extract(sequence_B, sequences_D) for sequence_B, sequences_D in zip(sequences_B, sequences_D)]

# Write the modified pegRNA protospacer sequences to the Excel file
for index, sequence in enumerate(modified_sequences_D):
    sheet[f'D{index+2}'].value = sequence

# Save the modified Excel file
output_file_path = "/Users/qichenyuan/Desktop/BootStrap Bio/NAFLD_03302024_modified.xlsx"
workbook.save(output_file_path)

print("Modification complete. Result saved to:", output_file_path)
